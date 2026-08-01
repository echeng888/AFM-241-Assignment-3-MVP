"""
AFM 241 Assignment 3 - MVP pipeline
Baker Tilly KDN | corporate investment bookkeeping

Four stages, matching section 6 of the write-up:
  1 ingestion              PDFs in, pages accounted for
  2 extraction             free text -> one validated record per transaction
  3 classification + math  model assigns treatment; CODE does every calculation
  4 output + flagging      five-tab workbook, two gates before a line reaches a total

RUN MODES
  --live     stage 2 and 3 call the Claude API. Requires ANTHROPIC_API_KEY.
  --offline  stage 2 uses a deterministic layout parser and stage 3 uses a
             keyword classifier, so the pipeline can be demonstrated end to end
             without an API key. Offline mode is NOT the product - it exists so
             stages 1 and 4 can be exercised and scored. The workbook records
             which mode produced it.

DATA
  Synthetic only. No real client statements were used at any point.

NOTE ON API DETAILS
  The live path below is written against the Anthropic Messages REST API using
  document blocks. Verify the current request shape, model names and field names
  against the vendor documentation before relying on it - do not take the syntax
  here as authoritative.
"""
import argparse, base64, json, os, re, subprocess, sys, datetime
from decimal import Decimal

OUT = "out"
MODEL = "claude-sonnet-4-6"

COMPLEX_PATTERNS = [
    (r"\bRET(?:URN)?\.?\s*OF\s*CAP", "Return of capital"),
    (r"\bROC\b", "Return of capital"),
    (r"NON-?TAXABLE\s+DIST", "Return of capital"),
    (r"PLAN OF ARRANGEMENT|MERGER|AMALGAMAT|REORGANIZ|EXCHANGED FOR", "Merger / reorganisation"),
    (r"REINVEST", "Distribution paid in shares"),
    (r"IN KIND|IN-KIND|DELIVERY OUT", "In-kind movement of securities"),
    (r"SHAREHOLDER", "Shareholder draw / benefit"),
    (r"SEE T3|SEE T5", "Character not determinable from the statement"),
]

CONFIDENCE_FLOOR = 0.80


# ---------------------------------------------------------------- stage 1
def ingest(statement_pdf, slips_pdf):
    def pages(p):
        n = subprocess.run(["pdfinfo", p], capture_output=True, text=True).stdout
        return int(re.search(r"Pages:\s+(\d+)", n).group(1))
    sp, sl = pages(statement_pdf), pages(slips_pdf)
    text = {}
    for i in range(1, sp + 1):
        t = subprocess.run(["pdftotext", "-layout", "-f", str(i), "-l", str(i), statement_pdf, "-"],
                           capture_output=True, text=True).stdout
        if not t.strip():
            raise SystemExit(f"STAGE 1 REJECT: page {i} of {statement_pdf} has no text layer "
                             f"(image-only or illegible scan). Nothing downstream will run.")
        text[i] = t
    print(f"[1] ingestion      {sp} statement pages + {sl} slip page(s), all with a text layer")
    return text, sp


# ---------------------------------------------------------------- stage 2
ROW = re.compile(
    r"^\s*(?P<date>\d{4}-\d{2}-\d{2})\s+(?P<rest>.+?)\s+(?P<qty>\(?-?[\d,]*\)?)\s+"
    r"(?P<amt>\(?[\d,]+\.\d{2}\)?)\s+(?P<ccy>[A-Z]{3})\s*$")

SECURITIES = ["Canadian Utility Trust Units", "Meridian Bank Common", "Northbridge Energy Corp",
              "Ridgeline Industrial Inc", "Halcyon Global Income Fund", "Atlas Materials Ltd",
              "Cascade Resources Ltd", "Fairview Semiconductor Inc (USD)", "Cash Balance"]


def _num(s):
    s = (s or "").strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "")
    if s in ("", "-"):
        return None
    v = float(s)
    return -v if neg else v


def extract_offline(page_text):
    """Deterministic layout parser. Stands in for the model so stages 1 and 4
    can be run and scored without an API key."""
    records = []
    for page, text in page_text.items():
        current = None
        for line in text.splitlines():
            m = ROW.match(line)
            if m:
                rest = m.group("rest").strip()
                sec = next((s for s in SECURITIES if rest.startswith(s)), None)
                if sec is None:
                    sec, desc = rest, rest
                else:
                    desc = rest[len(sec):].strip()
                current = {
                    "txn_id": f"T{len(records)+1:03d}",
                    "statement_month": page,
                    "source_page": page,
                    "txn_date": m.group("date"),
                    "security_name": sec,
                    "description_raw": desc,
                    "quantity": _num(m.group("qty")),
                    "gross_amount": _num(m.group("amt")),
                    "currency": m.group("ccy"),
                }
                records.append(current)
                continue
            # description continuation: an indented line with no date and no amount,
            # immediately following a transaction row
            if current is None:
                continue
            stripped = line.strip()
            if not stripped:
                current = None
                continue
            if re.match(r"^\d{4}-\d{2}-\d{2}", stripped) or "SYNTHETIC" in stripped:
                current = None
                continue
            current["description_raw"] = (current["description_raw"] + " " + stripped).strip()
    return records


LIVE_EXTRACT_PROMPT = """You are reading one page of a brokerage account statement.

Return ONLY a JSON array. One object per transaction row. No prose, no markdown,
no code fences. If the page has no transactions, return [].

Each object must have exactly these keys:
  txn_date         ISO date as printed, or null if you cannot read it
  security_name    as printed
  description_raw  the description column EXACTLY as printed, unmodified
  quantity         number or null
  gross_amount     number, negative for outflows, as printed
  currency         three-letter code

Do not interpret, normalise or correct anything. Do not compute. Do not infer a
value that is not printed on the page. If a field is unreadable, use null."""


def extract_live(statement_pdf, page_text):
    import urllib.request
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        raise SystemExit("--live needs ANTHROPIC_API_KEY set.")
    pdf_b64 = base64.b64encode(open(statement_pdf, "rb").read()).decode()
    records = []
    for page in sorted(page_text):
        body = {
            "model": MODEL,
            "max_tokens": 4000,
            "messages": [{"role": "user", "content": [
                {"type": "document",
                 "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                {"type": "text", "text": f"{LIVE_EXTRACT_PROMPT}\n\nRead page {page} only."},
            ]}],
        }
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages", method="POST",
            data=json.dumps(body).encode(),
            headers={"content-type": "application/json", "x-api-key": key,
                     "anthropic-version": "2023-06-01"})
        with urllib.request.urlopen(req, timeout=120) as r:
            data = json.loads(r.read())
        text = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
        text = text.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        try:
            rows = json.loads(text)
        except json.JSONDecodeError:
            raise SystemExit(f"STAGE 2 QUARANTINE: page {page} did not return valid JSON. "
                             f"The month is not processed on partial data.")
        for row in rows:
            row.update({"txn_id": f"T{len(records)+1:03d}", "statement_month": page,
                        "source_page": page})
            records.append(row)
    return records


REQUIRED = ["txn_id", "statement_month", "source_page", "txn_date", "security_name",
            "description_raw", "gross_amount", "currency"]


def validate(records):
    bad = [r for r in records if any(r.get(k) in (None, "") for k in REQUIRED)]
    if bad:
        months = sorted({r.get("statement_month") for r in bad})
        raise SystemExit(f"STAGE 2 QUARANTINE: {len(bad)} record(s) missing required fields "
                         f"in month(s) {months}. Pipeline stopped rather than proceeding "
                         f"on partial data.")
    return records


# ---------------------------------------------------------------- stage 3
KEYWORDS = [
    ("return_of_capital", r"RET(?:URN)?\.?\s*OF\s*CAP|\bROC\b|NON-?TAXABLE\s+DIST", 0.95),
    ("transfer",          r"PLAN OF ARRANGEMENT|EXCHANGED FOR|MERGER|AMALGAMAT", 0.60),
    ("withdrawal",        r"DELIVERY OUT|IN KIND|IN-KIND|FUNDS TRANSFER OUT|SHAREHOLDER", 0.70),
    ("dividend",          r"DIVIDEND|DIV\b|DIST\b|REINVEST", 0.90),
    ("interest",          r"INTEREST", 0.97),
    ("buy",               r"PURCHASE|\bBUY\b", 0.97),
    ("sell",              r"\bSELL\b|REDEMPTION", 0.96),
]


def classify_offline(records):
    for r in records:
        d = (r["description_raw"] or "").upper()
        r["txn_type"], r["confidence"], r["reason"] = "unknown", 0.30, "No keyword matched"
        for t, pat, conf in KEYWORDS:
            if re.search(pat, d):
                r["txn_type"], r["confidence"] = t, conf
                r["reason"] = f"Description matched the {t.replace('_',' ')} pattern"
                break
        if "SEE T3" in d or "SEE T5" in d:
            r["confidence"] = 0.45
            r["reason"] = "Statement defers the character of this amount to the slip"
    return records


LIVE_CLASSIFY_PROMPT = """You are classifying transactions from a Canadian corporate
investment account for tax working-paper purposes.

For each transaction below, return ONE JSON object with these keys:
  txn_id      echoed back
  txn_type    one of: buy, sell, dividend, interest, return_of_capital,
              contribution, withdrawal, transfer, fee, unknown
  confidence  0 to 1, your calibrated confidence in txn_type
  reason      one short sentence a reviewer would find useful

Return ONLY a JSON array, no prose and no code fences.

Rules:
- Do NOT calculate anything. Do not total, net or reconcile.
- "unknown" with low confidence is a correct and expected answer when the
  statement does not determine the treatment.
- Judge the meaning of the description, not its exact wording.

Transactions:
"""


def classify_live(records):
    import urllib.request
    key = os.environ["ANTHROPIC_API_KEY"]
    payload = [{"txn_id": r["txn_id"], "security_name": r["security_name"],
                "description_raw": r["description_raw"], "quantity": r.get("quantity"),
                "gross_amount": r["gross_amount"], "currency": r["currency"]} for r in records]
    body = {"model": MODEL, "max_tokens": 8000,
            "messages": [{"role": "user",
                          "content": LIVE_CLASSIFY_PROMPT + json.dumps(payload, indent=1)}]}
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", method="POST",
        data=json.dumps(body).encode(),
        headers={"content-type": "application/json", "x-api-key": key,
                 "anthropic-version": "2023-06-01"})
    with urllib.request.urlopen(req, timeout=180) as r:
        data = json.loads(r.read())
    text = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")
    text = text.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    by_id = {c["txn_id"]: c for c in json.loads(text)}
    for r in records:
        c = by_id.get(r["txn_id"], {})
        r["txn_type"] = c.get("txn_type", "unknown")
        r["confidence"] = float(c.get("confidence", 0.0))
        r["reason"] = c.get("reason", "Not returned by the classifier")
    return records


def apply_gates(records, reporting_ccy):
    """Two gates. Neither is a model decision."""
    for r in records:
        flags = []
        if r["confidence"] < CONFIDENCE_FLOOR:
            flags.append(f"Confidence {r['confidence']:.2f} below the {CONFIDENCE_FLOOR:.2f} floor")
        d = (r["description_raw"] or "").upper()
        for pat, label in COMPLEX_PATTERNS:
            if re.search(pat, d):
                flags.append(label)
        if r["currency"] != reporting_ccy:
            flags.append(f"{r['currency']} holding - cost base held in original currency at original date")
        r["flags"] = sorted(set(flags))
        r["in_totals"] = not r["flags"]
    return records


def compute(records, opening, slips, reporting_ccy):
    """Every number in the workbook is produced here, in code. The model
    contributed txn_type, confidence and reason - nothing else."""
    inc = {"dividend": 0.0, "interest": 0.0, "return_of_capital": 0.0}
    for r in records:
        if r["in_totals"] and r["txn_type"] in inc:
            inc[r["txn_type"]] += r["gross_amount"]

    positions = {p["security"]: dict(p) for p in opening["positions"]}
    activity = {}
    for r in records:
        sec = r["security_name"]
        if sec == "Cash Balance":
            continue
        a = activity.setdefault(sec, {"purchases": 0.0, "disposals": 0.0, "roc": 0.0, "other": 0.0})
        if not r["in_totals"]:
            a["other"] += 0.0
            continue
        if r["txn_type"] == "buy":
            a["purchases"] += -r["gross_amount"]
        elif r["txn_type"] == "sell":
            a["disposals"] += -r["gross_amount"]
        elif r["txn_type"] == "return_of_capital":
            a["roc"] += -r["gross_amount"]

    rollforward = []
    for sec in sorted(set(list(positions) + list(activity))):
        op = positions.get(sec, {"cost_base": 0.0, "currency": reporting_ccy, "quantity": 0})
        a = activity.get(sec, {"purchases": 0.0, "disposals": 0.0, "roc": 0.0})
        flagged = [r for r in records if r["security_name"] == sec and not r["in_totals"]]
        rollforward.append({
            "security": sec, "currency": op.get("currency", reporting_ccy),
            "opening": op.get("cost_base", 0.0),
            "purchases": a["purchases"], "disposals": a["disposals"], "roc": a["roc"],
            "closing": op.get("cost_base", 0.0) + a["purchases"] + a["disposals"] + a["roc"],
            "unresolved": len(flagged),
        })

    recon = [
        ("Eligible dividends", inc["dividend"], slips["T5_dividends_actual"], "T5 box 24"),
        ("Interest", inc["interest"], slips["T5_interest"], "T5 box 13"),
        ("Trust distributions", sum(r["gross_amount"] for r in records
                                    if r["in_totals"] and "Halcyon" in r["security_name"]),
         slips["T3_distributions"], "T3 total"),
    ]
    return rollforward, recon, inc


# ---------------------------------------------------------------- stage 4
def write_workbook(path, records, rollforward, recon, mode, meta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    HEAD = PatternFill("solid", fgColor="1F3B4D")
    FLAGF = PatternFill("solid", fgColor="F8E4E1")
    OKF = PatternFill("solid", fgColor="E8F1EB")
    WHITE = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    BASE = Font(name="Arial", size=9)
    BOLD = Font(name="Arial", size=9, bold=True)
    TITLE = Font(name="Arial", size=12, bold=True, color="1F3B4D")
    thin = Side(style="thin", color="D0D0D0")
    BOX = Border(bottom=thin)

    wb = Workbook()

    def head(ws, row, labels, widths):
        for i, (lab, w) in enumerate(zip(labels, widths), start=1):
            c = ws.cell(row=row, column=i, value=lab)
            c.fill, c.font = HEAD, WHITE
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    # ---- 1 Transactions
    ws = wb.active
    ws.title = "1 Transactions"
    ws["A1"] = "Transactions extracted from statements"
    ws["A1"].font = TITLE
    head(ws, 3, ["Txn ID", "Stmt month", "Src page", "Date", "Security",
                 "Description as printed", "Quantity", "Gross amount", "Curr"],
         [9, 11, 9, 12, 30, 42, 11, 14, 7])
    for i, r in enumerate(records, start=4):
        vals = [r["txn_id"], r["statement_month"], r["source_page"], r["txn_date"],
                r["security_name"], r["description_raw"], r.get("quantity"),
                r["gross_amount"], r["currency"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font, c.border = BASE, BOX
            if j == 8:
                c.number_format = '#,##0.00;(#,##0.00)'

    # ---- 2 Classification
    ws = wb.create_sheet("2 Classification")
    ws["A1"] = "Tax treatment per transaction"
    ws["A1"].font = TITLE
    ws["A2"] = "txn_type, confidence and reason are the model's output. No number on this tab was produced by the model."
    ws["A2"].font = Font(name="Arial", size=8, italic=True, color="666666")
    head(ws, 4, ["Txn ID", "Description as printed", "Tax treatment", "Confidence",
                 "Reason", "In totals?"], [9, 42, 20, 12, 46, 11])
    for i, r in enumerate(records, start=5):
        vals = [r["txn_id"], r["description_raw"], r["txn_type"], r["confidence"],
                r["reason"], "yes" if r["in_totals"] else "NO - flagged"]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font, c.border = BASE, BOX
            if j == 4:
                c.number_format = "0.00"
            if j == 6:
                c.fill = OKF if r["in_totals"] else FLAGF

    # ---- 3 Cost-base continuity
    ws = wb.create_sheet("3 Cost base continuity")
    ws["A1"] = "Cost-base continuity schedule"
    ws["A1"].font = TITLE
    ws["A2"] = "Opening + activity = closing. Flagged transactions are excluded from activity until resolved."
    ws["A2"].font = Font(name="Arial", size=8, italic=True, color="666666")
    head(ws, 4, ["Security", "Curr", "Opening cost base", "Purchases", "Disposals (cost)",
                 "Return of capital", "Closing cost base", "Unresolved items"],
         [32, 7, 17, 14, 16, 17, 17, 15])
    r0 = 5
    for i, rf in enumerate(rollforward):
        row = r0 + i
        ws.cell(row=row, column=1, value=rf["security"]).font = BASE
        ws.cell(row=row, column=2, value=rf["currency"]).font = BASE
        for col, key in [(3, "opening"), (4, "purchases"), (5, "disposals"), (6, "roc")]:
            c = ws.cell(row=row, column=col, value=round(rf[key], 2))
            c.font, c.number_format = BASE, '#,##0.00;(#,##0.00)'
        c = ws.cell(row=row, column=7, value=f"=SUM(C{row}:F{row})")
        c.font, c.number_format = BOLD, '#,##0.00;(#,##0.00)'
        c2 = ws.cell(row=row, column=8, value=rf["unresolved"])
        c2.font = BASE
        if rf["unresolved"]:
            c2.fill = FLAGF
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = BOX
    tot = r0 + len(rollforward)
    ws.cell(row=tot, column=1, value="Total").font = BOLD
    for col in "CDEFG":
        c = ws[f"{col}{tot}"]
        c.value = f"=SUM({col}{r0}:{col}{tot-1})"
        c.font, c.number_format = BOLD, '#,##0.00;(#,##0.00)'

    # ---- 4 Reconciliation
    ws = wb.create_sheet("4 Reconciliation")
    ws["A1"] = "Reconciliation to slips"
    ws["A1"].font = TITLE
    ws["A2"] = "A variance is stated, never absorbed. Amounts on flagged transactions are excluded from 'per statements'."
    ws["A2"].font = Font(name="Arial", size=8, italic=True, color="666666")
    head(ws, 4, ["Item", "Per statements", "Per slip", "Variance", "Source", "Status"],
         [28, 16, 14, 14, 16, 34])
    for i, (label, computed, slip, src) in enumerate(recon):
        row = 5 + i
        ws.cell(row=row, column=1, value=label).font = BASE
        for col, v in [(2, round(computed, 2)), (3, round(slip, 2))]:
            c = ws.cell(row=row, column=col, value=v)
            c.font, c.number_format = BASE, '#,##0.00;(#,##0.00)'
        c = ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        c.font, c.number_format = BOLD, '#,##0.00;(#,##0.00)'
        ws.cell(row=row, column=5, value=src).font = BASE
        diff = round(computed - slip, 2)
        stat = "Ties" if abs(diff) < 0.005 else f"VARIANCE {diff:,.2f} - investigate"
        c = ws.cell(row=row, column=6, value=stat)
        c.font, c.fill = BASE, (OKF if abs(diff) < 0.005 else FLAGF)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BOX

    # ---- 5 Exceptions
    ws = wb.create_sheet("5 Exceptions")
    ws["A1"] = "Exceptions - resolve before the schedule is signed"
    ws["A1"].font = TITLE
    ws["A2"] = ("Every transaction below is excluded from the totals until a preparer resolves it. "
                "A transaction reaches this tab if confidence is below the floor OR it matches a "
                "known-complex pattern - the second test is a rule, not a model decision.")
    ws["A2"].font = Font(name="Arial", size=8, italic=True, color="666666")
    head(ws, 4, ["Txn ID", "Date", "Src page", "Security", "Description as printed",
                 "Amount", "Model's read", "Conf.", "Why it is here", "What the preparer must determine"],
         [9, 12, 9, 28, 40, 13, 17, 8, 34, 42])
    ex = [r for r in records if not r["in_totals"]]
    ASK = {
        "Return of capital": "Confirm ROC per the trust's tax information; reduce ACB, do not report as income.",
        "Merger / reorganisation": "Confirm continuity of the holding and the cost base carried across; not a disposal.",
        "Distribution paid in shares": "Split into income and a new parcel; set that parcel's cost base at this date.",
        "In-kind movement of securities": "Determine deemed proceeds and whether a shareholder benefit arises.",
        "Shareholder draw / benefit": "Post to shareholder account, not portfolio activity.",
        "Character not determinable from the statement": "Take the character from the slip once received.",
    }
    for i, r in enumerate(ex, start=5):
        why = "; ".join(r["flags"])
        ask = next((ASK[k] for k in ASK if any(k in f for f in r["flags"])), "")
        if any("holding - cost base" in f for f in r["flags"]) and not ask:
            ask = "Hold the cost base in the original currency at the original acquisition date."
        vals = [r["txn_id"], r["txn_date"], r["source_page"], r["security_name"],
                r["description_raw"], r["gross_amount"], r["txn_type"], r["confidence"], why, ask]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font, c.border = BASE, BOX
            c.alignment = Alignment(vertical="top", wrap_text=(j in (5, 9, 10)))
            if j == 6:
                c.number_format = '#,##0.00;(#,##0.00)'
            if j == 8:
                c.number_format = "0.00"
            if j in (9, 10):
                c.fill = FLAGF

    # ---- 0 Run info
    ws = wb.create_sheet("0 Run info", 0)
    ws["A1"] = "Draft working paper - requires preparer and reviewer sign-off"
    ws["A1"].font = TITLE
    info = [
        ("Client", meta["client"]), ("Account", meta["account"]),
        ("Fiscal year", meta["fiscal_year"]), ("Reporting currency", meta["reporting_currency"]),
        ("Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Pipeline mode", mode),
        ("Transactions extracted", len(records)),
        ("Transactions in totals", sum(1 for r in records if r["in_totals"])),
        ("Transactions on exceptions tab", sum(1 for r in records if not r["in_totals"])),
        ("Confidence floor", CONFIDENCE_FLOOR),
        ("Data", "SYNTHETIC TEST DATA ONLY - no real client information was used"),
        ("Status", "DRAFT - not a filed return; every figure requires review"),
        ("Prepared by", ""), ("Reviewed by", ""),
    ]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 66
    for i, (k, v) in enumerate(info, start=3):
        a, b = ws.cell(row=i, column=1, value=k), ws.cell(row=i, column=2, value=v)
        a.font, b.font = BOLD, BASE
        a.border = b.border = BOX
    for w in wb.worksheets:
        w.page_setup.orientation = "landscape"
        w.page_setup.fitToWidth = 1
        w.page_setup.fitToHeight = 0
        w.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        w.print_options.horizontalCentered = True
    wb.save(path)


# ---------------------------------------------------------------- scoring
def score(records, gt_path):
    import csv
    gt = list(csv.DictReader(open(gt_path)))
    by_desc = {(g["txn_date"], g["security"], round(float(g["gross_amount"]), 2)): g for g in gt}
    matched = extracted_ok = type_ok = routine = 0
    seeded = [g for g in gt if g["must_be_flagged"] == "YES"]
    caught = 0
    false_flags = 0
    routine_total = 0
    desc_exact = 0
    for r in records:
        k = (r["txn_date"], r["security_name"], round(r["gross_amount"], 2))
        g = by_desc.get(k)
        if not g:
            continue
        matched += 1
        if abs(float(g["gross_amount"]) - r["gross_amount"]) < 0.005:
            extracted_ok += 1
        if g["description_as_printed"].strip() == (r["description_raw"] or "").strip():
            desc_exact += 1
        if g["must_be_flagged"] == "YES":
            if not r["in_totals"]:
                caught += 1
        else:
            routine_total += 1
            if r["txn_type"] == g["true_type"]:
                type_ok += 1
            if not r["in_totals"]:
                false_flags += 1
    return {
        "ground_truth_rows": len(gt), "extracted": len(records), "matched": matched,
        "extraction_accuracy": extracted_ok / len(gt),
        "description_verbatim_accuracy": desc_exact / len(gt),
        "classification_accuracy_routine": type_ok / routine_total if routine_total else 0,
        "seeded_cases": len(seeded), "seeded_caught": caught,
        "exception_recall": caught / len(seeded) if seeded else 1.0,
        "false_flag_rate": false_flags / routine_total if routine_total else 0,
    }


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--live", action="store_true", help="use the Claude API for stages 2 and 3")
    ap.add_argument("--statements", default=f"{OUT}/statements_FY2025.pdf")
    ap.add_argument("--slips", default=f"{OUT}/slips_FY2025.pdf")
    ap.add_argument("--opening", default=f"{OUT}/opening_cost_base.json")
    ap.add_argument("--slipvalues", default=f"{OUT}/slips.json")
    ap.add_argument("--truth", default=f"{OUT}/ground_truth.csv")
    ap.add_argument("--out", default=f"{OUT}/working_paper_FY2025.xlsx")
    a = ap.parse_args()

    mode = "LIVE - Claude API for extraction and classification" if a.live else \
           "OFFLINE - deterministic parser and keyword classifier (no API key present)"
    print(f"\nAFM 241 MVP | mode: {mode}\n" + "-" * 72)

    opening = json.load(open(a.opening))
    slips = json.load(open(a.slipvalues))
    ccy = opening["reporting_currency"]

    page_text, npages = ingest(a.statements, a.slips)

    records = extract_live(a.statements, page_text) if a.live else extract_offline(page_text)
    validate(records)
    print(f"[2] extraction     {len(records)} records, all schema-valid")

    records = classify_live(records) if a.live else classify_offline(records)
    records = apply_gates(records, ccy)
    rollforward, recon, inc = compute(records, opening, slips, ccy)
    flagged = sum(1 for r in records if not r["in_totals"])
    print(f"[3] classification {len(records)-flagged} into totals, {flagged} routed to exceptions")
    print(f"    every figure below this line was computed in code, not by the model")

    write_workbook(a.out, records, rollforward, recon, mode, opening)
    print(f"[4] output         {a.out}")

    s = score(records, a.truth)
    print("-" * 72)
    print("SCORED AGAINST THE HAND-BUILT GROUND TRUTH")
    print(f"  1 extraction accuracy            {s['extraction_accuracy']:6.1%}   target >= 95%   "
          f"{'PASS' if s['extraction_accuracy'] >= .95 else 'FAIL'}")
    print(f"    description captured verbatim        {s['description_verbatim_accuracy']:6.1%}")
    print(f"  2 classification, routine types  {s['classification_accuracy_routine']:6.1%}   target >= 90%   "
          f"{'PASS' if s['classification_accuracy_routine'] >= .90 else 'FAIL'}")
    v = [r for r in recon if abs(r[1] - r[2]) >= 0.005]
    print(f"  3 reconciliation                 {len(v)} variance(s) stated, 0 absorbed   PASS")
    print(f"  4 exception recall               {s['exception_recall']:6.1%}   target 100%     "
          f"{'PASS' if s['exception_recall'] >= 1.0 else 'FAIL'}  "
          f"({s['seeded_caught']}/{s['seeded_cases']} seeded cases)")
    print(f"  5 false-flag rate                {s['false_flag_rate']:6.1%}   ceiling 15%     "
          f"{'PASS' if s['false_flag_rate'] <= .15 else 'FAIL'}")
    print("-" * 72)
    json.dump(s, open(f"{OUT}/score.json", "w"), indent=2)


if __name__ == "__main__":
    main()
